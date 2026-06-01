from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def gerar_excel(candidatos: list[dict]) -> bytes:
    """Gera um .xlsx com os candidatos e retorna os bytes (BytesIO)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatos"

    cabecalho = ["Ano", "Municipio", "Cargo", "Nome", "Nome Urna", "Partido", "Situacao"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    for col, titulo in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    eleito_fill = PatternFill("solid", fgColor="C6EFCE")

    for i, c in enumerate(candidatos, 2):
        valores = [c["ano"], c["municipio"], c["cargo"], c["nome"],
                   c["urna"], c["partido"], c["situacao"]]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if c["eleito"]:
                cell.fill = eleito_fill

    larguras = [6, 20, 12, 40, 25, 10, 20]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = larg

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
